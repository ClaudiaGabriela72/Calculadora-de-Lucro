import openpyxl


#nome da folha excel
nome = 'planilha.xlsx'

#obetendo os dados
def obter_dados_excel(i):
    wb = openpyxl.load_workbook(i)
    sheet = wb.active
    dados = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        dados.append(row)

    return dados


#Deletar produto
def deletar_linha_por_nome(nome_produto, nome_planilha):
    wb = openpyxl.load_workbook(nome_planilha)
    sheet = wb.active
    contador = 2

    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only= True):
        if str([0]) == nome_produto:
            # Obter o numero da linha e deletar a linha interira
            linha = contador
            sheet.delete_rows(linha)
            break